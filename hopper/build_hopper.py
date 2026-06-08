"""
Build Peter Tosh's Fit x Signal triage + machine-readable hopper for the
self-driving Apollo pipeline. Same scoring method and xlsx format as the
Jason Patch triage (account-scoring skill, Fit x Signal v1, light public-data pass).

Scoring grounded in:
- account-scoring-skill/references/scoring-rules.md (named anchors for all 11)
- 2026-05-29 Top 10 dated signals (capex, pain, conference, auto-moves, exec churn)
- On-disk research files (C&S, Cintas, Honda) raise confidence

Outputs:
- Peter_Patch_FitSignal_Triage_<date>.xlsx  (human leaderboard, 4 sheets)
- hopper.jsonl                              (machine queue the Morning Driver reads)
"""

import json
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Gather brand
NAVY = "0A2540"
GOLD = "C9A227"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
STRIKE_COLOR = "1E7E34"
NURTURE_COLOR = "FFC107"
QUALIFY_COLOR = "17A2B8"
KILL_COLOR = "C82333"
ARIAL = "Arial"

HERE = "/Users/petertosh/Desktop/Claude Cowork/OUTPUTS/Apollo Pipeline Orchestrator/hopper"

# ----- DATA -----
# fit  = [facility, vertical, throughput, automation]
# sig  = [capex, public_pain, conference, auto_moves, exec_churn]
accounts = [
    {
        "name": "Walmart", "domain": "walmart.com",
        "fit": [10, 10, 10, 9], "sig": [9, 8, 7, 9, 8], "conf": "H",
        "anchor": "~210 US DCs, 50k+ pallets/day peak, conventional non-HRDC fleet. Symbotic multi-site rollout but 19 non-Symbotic RDCs still open for forklift-side vision.",
        "move": "Cedric Clark Store Ops seat opens in weeks. Queue the intro for the day the name drops. Open on Prepaid Consolidation (live 5/26 across 42 RDCs).",
    },
    {
        "name": "DHL Supply Chain", "domain": "dhl.com",
        "fit": [10, 9, 9, 8], "sig": [9, 6, 6, 9, 9], "conf": "H",
        "anchor": "Global top-tier 3PL. SOFTBOT + Locus + Boston Dynamics across 100+ sites, 8,000 robots. Annville 1M sqft FDA/GMP DC opening into that stack.",
        "move": "Three new seats on one announcement (Monkmeyer, Ruff, Moss). Sharp one-pager to Monkmeyer this week. Tee Ruff up on the Annville pharma DC.",
    },
    {
        "name": "Sysco", "domain": "sysco.com",
        "fit": [9, 9, 8, 7], "sig": [7, 5, 9, 7, 8], "conf": "H",
        "anchor": "80+ broadline + specialty sites, foodservice core ICP. Jetro $29.1B integration with $250M cost-out. SAGE just won an AI Impact Award.",
        "move": "Re-engage interim CFO Sewell and interim CIO Advani before permanent hires. POV: vision on the order picker, no WMS replatform.",
    },
    {
        "name": "C&S Wholesale Grocers", "domain": "cswg.com",
        "fit": [9, 9, 9, 7], "sig": [6, 5, 5, 7, 8], "conf": "H",
        "anchor": "Largest US grocery wholesaler, 50+ already-automated DCs, mid-migration to GCP for AI forecasting. Symbotic handles case throughput at AWG.",
        "move": "Papaleo retiring as EVP/CPO, Gross to co-president. Use Gross promotion as warm note, ask connect to incoming CPO. Operator-side visibility is the day-one gap.",
    },
    {
        "name": "P&G", "domain": "pg.com",
        "fit": [8, 8, 9, 8], "sig": [9, 5, 5, 7, 4], "conf": "H",
        "anchor": "Lima / Cape Girardeau high-throughput plants. Supply Chain 3.0 scaled rollout, $1.5B COGS target, $205M Georgia automated DC. Berlin dark-shift template.",
        "move": "Use Berlin unattended night production as the wedge into Jejurikar's team. Pitch audit/observability for the Georgia DC ramp before robotics goes live.",
    },
    {
        "name": "Honda", "domain": "honda.com",
        "fit": [8, 7, 7, 9], "sig": [5, 7, 4, 8, 7], "conf": "H",
        "anchor": "HTM service parts at conventional MHE scale, modern WMS, piloting Vector/Verity-class drones. Research file on disk. Parts recall scar tissue.",
        "move": "Thread the new HTM logistics leader. POV on parts-pick accuracy across the HTM network. Reuse the research file already on disk.",
    },
    {
        "name": "Ryder", "domain": "ryder.com",
        "fit": [9, 8, 7, 7], "sig": [5, 4, 5, 5, 4], "conf": "L",
        "anchor": "Supply Chain Solutions runs a large contract-warehouse network (3PL core ICP). Thin public signal in the current window. Needs a fresh research pass.",
        "move": "Confirm warehouse footprint scale, single-thread one Ops Director. Run account-research before this reaches the top of the queue.",
    },
    {
        "name": "Toyota (Motor / TMH)", "domain": "toyota.com",
        "fit": [8, 7, 7, 7], "sig": [6, 4, 5, 7, 6], "conf": "M",
        "anchor": "Auto manufacturing + parts distribution. TMH shipped two automated forklifts (Center-Controlled Rider, Core Tow Tractor) under Aaron Jones at TAL. Indiana plant +85 jobs.",
        "move": "Disambiguate Toyota Motor distribution vs TMH first. Thread the perception layer above TAL workflows (Bastian, Vanderlande, viastore, Raymond).",
    },
    {
        "name": "Cintas", "domain": "cintas.com",
        "fit": [8, 7, 7, 7], "sig": [5, 4, 7, 7, 5], "conf": "H",
        "anchor": "Route-distribution + rental footprint, conventional MHE. UniFirst shareholder vote 2026-06-11. Schneider $375M 4-year cost-out rationalizing combined garment routing. Research file on disk.",
        "move": "Integration-playbook one-pager loaded for 2026-06-12 send, win or lose. Don't queue Rozakis until after the vote. Schneider's integration video is the open.",
    },
    {
        "name": "GE (Vernova / Aerospace)", "domain": "ge.com",
        "fit": [6, 5, 5, 6], "sig": [7, 4, 4, 7, 7], "conf": "M",
        "anchor": "Diversified industrial, fit is sideways. Vernova bought Robotech Automation (5/21, Q3 close) adding integration capability inside Gas Power, no prior automation owner. On-disk work is GE Aerospace.",
        "move": "Get on the calendar inside 30 days, before integration freezes. Disambiguate Vernova vs Aerospace. Position vision on the forklift fleet at Cincinnati/Asheville builds.",
    },
    {
        "name": "City Furniture", "domain": "cityfurniture.com",
        "fit": [7, 6, 5, 6], "sig": [5, 3, 4, 5, 4], "conf": "L",
        "anchor": "~30 sites incl showrooms, single-DC ~3k pallets/day, FL retail/furniture distribution. Low public signal. Needs a fresh research pass.",
        "move": "Park; build a POV on FL DC accuracy. Run account-research before working. Confirm throughput before spending pamphlet effort.",
    },
]

# ----- COMPUTE -----
def axis(scores):
    return round(sum(scores) / len(scores) * 10, 1)

def verdict(fit, sig):
    if fit < 40 or sig < 40:
        return "KILL"
    if fit >= 70 and sig >= 70:
        return "STRIKE"
    if fit >= 70 and sig < 70:
        return "NURTURE"
    return "QUALIFY"

for a in accounts:
    a["fit_score"] = axis(a["fit"])
    a["sig_score"] = axis(a["sig"])
    a["verdict"] = verdict(a["fit_score"], a["sig_score"])

order_rank = {"STRIKE": 0, "NURTURE": 1, "QUALIFY": 2, "KILL": 3}
accounts.sort(key=lambda a: (order_rank[a["verdict"]], -(a["fit_score"] + a["sig_score"])))

# ----- WRITE hopper.jsonl (machine queue) -----
today = date.today().isoformat()
hopper_path = f"{HERE}/hopper.jsonl"
with open(hopper_path, "w") as f:
    for rank, a in enumerate(accounts, 1):
        rec = {
            "rank": rank,
            "account": a["name"],
            "domain": a["domain"],
            "verdict": a["verdict"],
            "fit": a["fit_score"],
            "signal": a["sig_score"],
            "confidence": a["conf"],
            "top_move": a["move"],
            "status": "todo",
            "last_worked_at": None,
            "sequence_target": None,
            "scored_at": today,
        }
        f.write(json.dumps(rec) + "\n")

# ----- BUILD WORKBOOK -----
wb = Workbook()
thin = Side(border_style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

verdict_colors = {"STRIKE": STRIKE_COLOR, "NURTURE": NURTURE_COLOR, "QUALIFY": QUALIFY_COLOR, "KILL": KILL_COLOR}
verdict_text = {"STRIKE": WHITE, "NURTURE": "1A1A1A", "QUALIFY": WHITE, "KILL": WHITE}

# Sheet 1: Leaderboard
ws = wb.active
ws.title = "Leaderboard"
ws.merge_cells("A1:H1")
ws["A1"] = "Peter's Patch — Fit × Signal Triage"
ws["A1"].font = Font(name=ARIAL, size=18, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", start_color=NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:H2")
ws["A2"] = f"11 accounts • Scored {today} • Light public-data pass • Source: account-scoring skill (Fit×Signal v1) + 2026-05-29 Top 10 signals"
ws["A2"].font = Font(name=ARIAL, size=10, italic=True, color="555555")
ws["A2"].alignment = Alignment(horizontal="left", indent=1)
ws.row_dimensions[2].height = 20

headers = ["#", "Account", "Verdict", "Fit /100", "Signal /100", "Conf.", "Why it scores this way", "Top move for Peter"]
ws.append([])
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = Font(name=ARIAL, bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[4].height = 30

for i, a in enumerate(accounts, 1):
    row = 4 + i
    ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=2, value=a["name"]).font = Font(name=ARIAL, bold=True, size=11)
    v = ws.cell(row=row, column=3, value=a["verdict"])
    v.fill = PatternFill("solid", start_color=verdict_colors[a["verdict"]])
    v.font = Font(name=ARIAL, bold=True, color=verdict_text[a["verdict"]], size=10)
    v.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4, value=a["fit_score"]).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=5, value=a["sig_score"]).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=6, value=a["conf"]).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=7, value=a["anchor"]).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=row, column=8, value=a["move"]).alignment = Alignment(wrap_text=True, vertical="top")
    if i % 2 == 0:
        for col in [1, 2, 4, 5, 6, 7, 8]:
            ws.cell(row=row, column=col).fill = PatternFill("solid", start_color=LIGHT_GRAY)
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = border
        if col not in (2, 3):
            ws.cell(row=row, column=col).font = Font(name=ARIAL, size=10)
    ws.row_dimensions[row].height = 60

widths = {"A": 5, "B": 26, "C": 12, "D": 10, "E": 12, "F": 8, "G": 56, "H": 56}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

last = 4 + len(accounts) + 2
ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=8)
ws.cell(row=last, column=1, value=(
    "Verdict math: Fit×Signal matrix. STRIKE = Fit≥70 AND Signal≥70. NURTURE = Fit≥70, Signal<70. "
    "QUALIFY = Fit 40-69, Signal≥40. KILL = either axis <40. Recalibrate quarterly or on trigger event. "
    "L-confidence rows (Ryder, City Furniture) need a fresh account-research pass before they reach the top of the work queue."
))
ws.cell(row=last, column=1).font = Font(name=ARIAL, size=9, italic=True, color="555555")
ws.cell(row=last, column=1).alignment = Alignment(wrap_text=True)
ws.freeze_panes = "B5"

# Sheet 2: Dimension Detail
ws2 = wb.create_sheet("Dimension Detail")
ws2.merge_cells("A1:O1")
ws2["A1"] = "9-Dimension Detail — Fit (4) × Signal (5)"
ws2["A1"].font = Font(name=ARIAL, size=16, bold=True, color=WHITE)
ws2["A1"].fill = PatternFill("solid", start_color=NAVY)
ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.row_dimensions[1].height = 30
dim_headers = ["Account", "Verdict", "Fit: Facilities", "Fit: Vertical", "Fit: Throughput", "Fit: Auto Maturity", "FIT /100",
    "Sig: Capex", "Sig: Public Pain", "Sig: Conference", "Sig: Auto Moves", "Sig: Exec Churn", "SIGNAL /100", "Confidence", "Notes"]
for col, h in enumerate(dim_headers, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = Font(name=ARIAL, bold=True, color=WHITE, size=9)
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws2.row_dimensions[3].height = 38
for i, a in enumerate(accounts, 1):
    row = 3 + i
    ws2.cell(row=row, column=1, value=a["name"]).font = Font(name=ARIAL, bold=True, size=10)
    v = ws2.cell(row=row, column=2, value=a["verdict"])
    v.fill = PatternFill("solid", start_color=verdict_colors[a["verdict"]])
    v.font = Font(name=ARIAL, bold=True, color=verdict_text[a["verdict"]], size=9)
    v.alignment = Alignment(horizontal="center", vertical="center")
    for j, val in enumerate(a["fit"]):
        ws2.cell(row=row, column=3 + j, value=val).alignment = Alignment(horizontal="center")
    fit_letters = [get_column_letter(3 + j) for j in range(4)]
    ws2.cell(row=row, column=7, value=f"=AVERAGE({fit_letters[0]}{row}:{fit_letters[3]}{row})*10").alignment = Alignment(horizontal="center")
    ws2.cell(row=row, column=7).font = Font(name=ARIAL, bold=True, size=10)
    for j, val in enumerate(a["sig"]):
        ws2.cell(row=row, column=8 + j, value=val).alignment = Alignment(horizontal="center")
    sig_letters = [get_column_letter(8 + j) for j in range(5)]
    ws2.cell(row=row, column=13, value=f"=AVERAGE({sig_letters[0]}{row}:{sig_letters[4]}{row})*10").alignment = Alignment(horizontal="center")
    ws2.cell(row=row, column=13).font = Font(name=ARIAL, bold=True, size=10)
    ws2.cell(row=row, column=14, value=a["conf"]).alignment = Alignment(horizontal="center")
    ws2.cell(row=row, column=15, value=a["anchor"]).alignment = Alignment(wrap_text=True, vertical="top")
    if i % 2 == 0:
        for col in range(1, 16):
            if col != 2:
                ws2.cell(row=row, column=col).fill = PatternFill("solid", start_color=LIGHT_GRAY)
    for col in range(1, 16):
        ws2.cell(row=row, column=col).border = border
        if col not in (1, 2, 7, 13, 15):
            ws2.cell(row=row, column=col).font = Font(name=ARIAL, size=9)
    ws2.row_dimensions[row].height = 54
w2 = {"A": 24, "B": 11, "C": 10, "D": 10, "E": 11, "F": 13, "G": 9, "H": 10, "I": 12, "J": 11, "K": 12, "L": 12, "M": 11, "N": 10, "O": 52}
for col, w in w2.items():
    ws2.column_dimensions[col].width = w
cs_rule = ColorScaleRule(start_type='num', start_value=1, start_color='F8D7DA', mid_type='num', mid_value=5, mid_color='FFF3CD', end_type='num', end_value=10, end_color='D4EDDA')
ws2.conditional_formatting.add(f"C4:F{3 + len(accounts)}", cs_rule)
ws2.conditional_formatting.add(f"H4:L{3 + len(accounts)}", cs_rule)
ws2.freeze_panes = "B4"

# Sheet 3: Quadrant Map
ws3 = wb.create_sheet("Quadrant Map")
ws3.merge_cells("A1:F1")
ws3["A1"] = "Fit × Signal Quadrant View"
ws3["A1"].font = Font(name=ARIAL, size=16, bold=True, color=WHITE)
ws3["A1"].fill = PatternFill("solid", start_color=NAVY)
ws3["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.row_dimensions[1].height = 30
quad_data = {
    "STRIKE\nFit ≥ 70, Signal ≥ 70": [a for a in accounts if a["verdict"] == "STRIKE"],
    "NURTURE\nFit ≥ 70, Signal < 70": [a for a in accounts if a["verdict"] == "NURTURE"],
    "QUALIFY\nFit 40-69, Signal ≥ 40": [a for a in accounts if a["verdict"] == "QUALIFY"],
    "KILL / IGNORE\nEither axis < 40": [a for a in accounts if a["verdict"] == "KILL"],
}
quad_colors = [STRIKE_COLOR, NURTURE_COLOR, QUALIFY_COLOR, KILL_COLOR]
quad_text_colors = [WHITE, "1A1A1A", WHITE, WHITE]
row_cursor = 3
for idx, (qname, items) in enumerate(quad_data.items()):
    ws3.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=6)
    h = ws3.cell(row=row_cursor, column=1, value=qname)
    h.font = Font(name=ARIAL, bold=True, size=12, color=quad_text_colors[idx])
    h.fill = PatternFill("solid", start_color=quad_colors[idx])
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws3.row_dimensions[row_cursor].height = 36
    row_cursor += 1
    if not items:
        ws3.cell(row=row_cursor, column=1, value="(none)").font = Font(name=ARIAL, italic=True, size=10, color="888888")
        row_cursor += 1
    else:
        for col, sh in enumerate(["Account", "Fit", "Signal", "Conf.", "Top move", ""], 1):
            c = ws3.cell(row=row_cursor, column=col, value=sh if sh else None)
            c.font = Font(name=ARIAL, bold=True, size=9, color=NAVY)
            c.border = border
        ws3.row_dimensions[row_cursor].height = 20
        row_cursor += 1
        for a in items:
            ws3.cell(row=row_cursor, column=1, value=a["name"]).font = Font(name=ARIAL, bold=True, size=10)
            ws3.cell(row=row_cursor, column=2, value=a["fit_score"]).alignment = Alignment(horizontal="center")
            ws3.cell(row=row_cursor, column=3, value=a["sig_score"]).alignment = Alignment(horizontal="center")
            ws3.cell(row=row_cursor, column=4, value=a["conf"]).alignment = Alignment(horizontal="center")
            ws3.merge_cells(start_row=row_cursor, start_column=5, end_row=row_cursor, end_column=6)
            ws3.cell(row=row_cursor, column=5, value=a["move"]).alignment = Alignment(wrap_text=True, vertical="top")
            for col in range(1, 7):
                ws3.cell(row=row_cursor, column=col).border = border
                if col not in (1, 5):
                    ws3.cell(row=row_cursor, column=col).font = Font(name=ARIAL, size=9)
            ws3.row_dimensions[row_cursor].height = 48
            row_cursor += 1
    row_cursor += 1
w3 = {"A": 26, "B": 8, "C": 10, "D": 8, "E": 40, "F": 30}
for col, w in w3.items():
    ws3.column_dimensions[col].width = w

# Sheet 4: Scoring Method
ws4 = wb.create_sheet("Scoring Method")
ws4.merge_cells("A1:B1")
ws4["A1"] = "Scoring Methodology"
ws4["A1"].font = Font(name=ARIAL, size=16, bold=True, color=WHITE)
ws4["A1"].fill = PatternFill("solid", start_color=NAVY)
ws4["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws4.row_dimensions[1].height = 30
methodology = [
    ("Source", "Gather AI Lead Workflow -> account-scoring skill v1, plus 2026-05-29 Top 10 dated signals."),
    ("Axes", "Fit (4 dims, avg x10) and Signal (5 dims, avg x10). Both /100."),
    ("Fit dims", "Facility count, Industry vertical, Throughput proxies, Automation maturity."),
    ("Signal dims", "Capex/builds, Public pain, Conference presence, Automation moves, C-suite/VP Ops churn."),
    ("Verdict logic", "STRIKE = Fit>=70 AND Signal>=70. NURTURE = Fit>=70, Signal<70. QUALIFY = Fit 40-69, Signal>=40. KILL = either axis <40."),
    ("Confidence", "H = research file on disk or strong public anchors. M = some anchors, some gaps. L = thin public data, directional only."),
    ("Data depth", "Light public-data pass. L-confidence accounts (Ryder, City Furniture) need a fresh account-research file before working."),
    ("Hopper", "Machine queue at hopper.jsonl drives the Morning Driver. STRIKE feeds first, top-down. Re-score on trigger event or quarterly."),
]
for i, (k, v) in enumerate(methodology):
    r = 3 + i
    ws4.cell(row=r, column=1, value=k).font = Font(name=ARIAL, bold=True, size=11, color=NAVY)
    ws4.cell(row=r, column=1).alignment = Alignment(vertical="top")
    ws4.cell(row=r, column=2, value=v).font = Font(name=ARIAL, size=10)
    ws4.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[r].height = 40
ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 90

xlsx_path = f"{HERE}/Peter_Patch_FitSignal_Triage_{today}.xlsx"
wb.save(xlsx_path)
print(f"Saved leaderboard: {xlsx_path}")
print(f"Saved hopper:      {hopper_path}")
print("\n=== VERDICT BREAKDOWN ===")
for v in ["STRIKE", "NURTURE", "QUALIFY", "KILL"]:
    matches = [f"{a['name']} ({a['fit_score']}/{a['sig_score']})" for a in accounts if a["verdict"] == v]
    print(f"{v} ({len(matches)}): {', '.join(matches) if matches else '-'}")
